#!/usr/bin/env python3
"""
Build the PR retrospective Excel workbook.

Usage:
  generate_report.py \
      --pr 6 \
      --title "Amazon MCF order fulfillment pipeline" \
      --base main --head mcf-order-fulfillment \
      --window 2026-03-31 2026-04-19 \
      --commits commits.json \
      --docs-audit docs-audit.json \
      --ci-timeline ci-timeline.json \
      --sessions sessions.json \
      --walkthrough-fixes walkthrough-fixes.json \
      --friction friction.json \
      --refinements refinements.json \
      --out retrospective-pr6.xlsx

Every tab has a fixed schema. If the corresponding input JSON is missing or
empty, the tab is still created with a "(no data)" note so downstream readers
see a consistent shape.

Input JSON shapes — see docstrings on each loader below.
"""

from __future__ import annotations

import argparse
import json
import re
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# openpyxl rejects control chars (except \t, \n, \r). Strip anything else.
_ILLEGAL_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def clean(value: Any) -> Any:
    if isinstance(value, str):
        return _ILLEGAL_RE.sub("", value)
    return value


HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SUBTLE_FILL = PatternFill("solid", fgColor="E7EEF7")
WRAP = Alignment(wrap_text=True, vertical="top")


def load(path: str | None) -> Any:
    if not path:
        return None
    p = Path(path)
    if not p.exists():
        return None
    try:
        return json.loads(p.read_text())
    except json.JSONDecodeError:
        return None


def write_header(ws: Worksheet, headers: list[str]) -> None:
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=clean(h))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = "A2"


def autosize(ws: Worksheet, max_width: int = 60) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        width = 10
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            val = row[0]
            if val is None:
                continue
            s = str(val)
            # Multi-line cells: size to longest line
            longest = max((len(x) for x in s.split("\n")), default=0)
            if longest > width:
                width = min(longest + 2, max_width)
        ws.column_dimensions[letter].width = width


def wrap_column(ws: Worksheet, col_idx: int) -> None:
    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
        for cell in row:
            cell.alignment = WRAP


def add_no_data(ws: Worksheet, headers: list[str]) -> None:
    ws.cell(row=2, column=1, value="(no data)").font = Font(italic=True, color="888888")
    for c in range(2, len(headers) + 1):
        ws.cell(row=2, column=c, value=clean(""))


_CI_FIX_RE = re.compile(r"^(fix\(ci\)|ci[:(]|ci\s)", re.IGNORECASE)


def parse_dt(s: str | None) -> datetime | None:
    """Parse ISO 8601 or `git log --date=iso` (space-separated) formats."""
    if not s or not isinstance(s, str):
        return None
    s = s.strip()
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00"))
    except ValueError:
        pass
    # git log format: "2026-03-31 16:23:27 -0700"
    try:
        return datetime.fromisoformat(s.replace(" ", "T", 1))
    except ValueError:
        return None


def enrich_commits(
    commits: list[dict],
    walkthrough_fixes: list[dict],
    ci_timeline: dict,
    override_rfr: str | None,
) -> tuple[list[dict], datetime | None]:
    """Infer is_walkthrough_fix / is_ci_fix / post_ready_for_review per commit.

    Respects pre-existing flags on input (manual judgment wins). Returns enriched
    list + the resolved ready-for-review timestamp.
    """
    walkthrough_shas: set[str] = set()
    for f in walkthrough_fixes or []:
        sha = (f.get("source_commit") or "").lower()
        if sha:
            walkthrough_shas.add(sha[:10])

    rfr_raw = override_rfr or (ci_timeline or {}).get("ready_for_review_ts")
    if not rfr_raw:
        # Infer: first commit whose message starts with a CI-fix pattern.
        for c in commits:
            if _CI_FIX_RE.match(c.get("message", "") or ""):
                rfr_raw = c.get("date", "")
                break
    rfr_dt = parse_dt(rfr_raw) if rfr_raw else None

    # Enrichment is authoritative when it finds a positive signal. An existing
    # True on the commit (e.g. from a manual override) is preserved (OR semantics).
    for c in commits:
        sha10 = (c.get("sha") or "").lower()[:10]
        c["is_walkthrough_fix"] = bool(c.get("is_walkthrough_fix")) or (sha10 in walkthrough_shas)
        c["is_ci_fix"] = bool(c.get("is_ci_fix")) or bool(
            _CI_FIX_RE.match(c.get("message", "") or "")
        )
        if rfr_dt:
            cdt = parse_dt(c.get("date"))
            if cdt and cdt.tzinfo is not None and rfr_dt.tzinfo is not None:
                inferred = cdt >= rfr_dt
            elif cdt and rfr_dt:
                inferred = cdt.replace(tzinfo=None) >= rfr_dt.replace(tzinfo=None)
            else:
                inferred = False
            c["post_ready_for_review"] = bool(c.get("post_ready_for_review")) or inferred
    return commits, rfr_dt


# ---------------------------------------------------------------------------
# Tab builders
# ---------------------------------------------------------------------------


def tab_summary(wb: Workbook, args: argparse.Namespace, data: dict) -> None:
    ws = wb.create_sheet("Summary")
    headers = ["Field", "Value"]
    write_header(ws, headers)

    sessions = (data.get("sessions") or {}).get("sessions") or []
    commits = data.get("commits") or []
    ci = data.get("ci_timeline") or {}
    docs = data.get("docs_audit") or {}

    rows = [
        ("PR number", f"#{args.pr}"),
        ("Title", args.title),
        ("Base branch", args.base),
        ("Head branch", args.head),
        ("Window start", args.window[0]),
        ("Window end", args.window[1]),
        ("Duration (days)", _days_between(args.window[0], args.window[1])),
        ("Total commits", len(commits)),
        ("Fix(ci) commits", sum(1 for c in commits if (c.get("type") or "").startswith("ci"))),
        ("Walkthrough fix commits", sum(1 for c in commits if c.get("is_walkthrough_fix"))),
        ("Net LOC (+/-)",
            f"+{sum(c.get('additions', 0) for c in commits)} "
            f"-{sum(c.get('deletions', 0) for c in commits)}"),
        ("Files changed", len({f for c in commits for f in (c.get('files') or [])})),
        ("Claude Code sessions in window", len(sessions)),
        ("Total messages across sessions", sum(s.get("message_count", 0) for s in sessions)),
        ("Compacted sessions",
            f"{sum(1 for s in sessions if s.get('compacted'))} of {len(sessions)}"),
        ("PR-prep doc files", len(docs.get("files") or [])),
        ("Specs written during window", len((data.get("specs") or {}).get("files") or [])),
        ("CI green first reached", ci.get("first_green_ts") or "(unknown)"),
        ("CI reruns after first ready-for-review", ci.get("reruns_after_ready") or 0),
    ]
    for r, (k, v) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=clean(k)).font = Font(bold=True)
        ws.cell(row=r, column=2, value=clean(v))
    autosize(ws, max_width=80)


def _days_between(a: str, b: str) -> int | str:
    try:
        from datetime import date
        da = date.fromisoformat(a[:10])
        db = date.fromisoformat(b[:10])
        return (db - da).days
    except ValueError:
        return "?"


def tab_timeline(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Timeline")
    headers = ["When", "Kind", "Ref", "Summary"]
    write_header(ws, headers)
    events: list[tuple[str, str, str, str]] = []
    for c in data.get("commits") or []:
        events.append((c.get("date", ""), "commit", c.get("sha", "")[:8], c.get("message", "")))
    for s in (data.get("sessions") or {}).get("sessions") or []:
        events.append((s.get("first_ts", ""), "session-start", s.get("id", "")[:8],
                       s.get("summary") or s.get("opening_prompt", "")))
    for e in (data.get("ci_timeline") or {}).get("events") or []:
        events.append((e.get("ts", ""), e.get("kind", "ci"), e.get("ref", ""),
                       e.get("summary", "")))
    if not events:
        add_no_data(ws, headers)
        autosize(ws)
        return
    events.sort(key=lambda x: x[0])
    for r, (when, kind, ref, summary) in enumerate(events, start=2):
        ws.cell(row=r, column=1, value=clean(when))
        ws.cell(row=r, column=2, value=clean(kind))
        ws.cell(row=r, column=3, value=clean(ref))
        ws.cell(row=r, column=4, value=clean(summary))
    wrap_column(ws, 4)
    autosize(ws, max_width=90)


def tab_commits(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Commits")
    headers = ["SHA", "Date", "Author", "Type", "Message", "Additions", "Deletions",
               "Files", "Walkthrough fix?", "CI fix?", "Post ready-for-review?"]
    write_header(ws, headers)
    commits = data.get("commits") or []
    if not commits:
        add_no_data(ws, headers)
        autosize(ws)
        return
    for r, c in enumerate(commits, start=2):
        ws.cell(row=r, column=1, value=clean((c.get("sha") or "")[:10]))
        ws.cell(row=r, column=2, value=clean(c.get("date", "")))
        ws.cell(row=r, column=3, value=clean(c.get("author", "")))
        ws.cell(row=r, column=4, value=clean(c.get("type", "")))
        ws.cell(row=r, column=5, value=clean(c.get("message", "")))
        ws.cell(row=r, column=6, value=clean(c.get("additions", 0)))
        ws.cell(row=r, column=7, value=clean(c.get("deletions", 0)))
        files = c.get("files") or []
        ws.cell(row=r, column=8, value=clean(", ".join(files[:5])) + (
            f" +{len(files) - 5} more" if len(files) > 5 else ""))
        ws.cell(row=r, column=9, value="Y" if c.get("is_walkthrough_fix") else "")
        ws.cell(row=r, column=10, value="Y" if c.get("is_ci_fix") else "")
        ws.cell(row=r, column=11, value="Y" if c.get("post_ready_for_review") else "")
    wrap_column(ws, 5)
    wrap_column(ws, 8)
    autosize(ws, max_width=70)


def tab_walkthrough_fixes(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Walkthrough Fixes")
    headers = ["#", "Description", "Category", "Blocker?", "Could defer?",
               "Notes / minimum-softening", "Source commit"]
    write_header(ws, headers)
    fixes = data.get("walkthrough_fixes") or []
    if not fixes:
        add_no_data(ws, headers)
        autosize(ws)
        return
    for r, f in enumerate(fixes, start=2):
        ws.cell(row=r, column=1, value=clean(f.get("index")) or r - 1)
        ws.cell(row=r, column=2, value=clean(f.get("description", "")))
        ws.cell(row=r, column=3, value=clean(f.get("category", "")))
        ws.cell(row=r, column=4, value=clean(f.get("blocker", "")))
        ws.cell(row=r, column=5, value=clean(f.get("could_defer", "")))
        ws.cell(row=r, column=6, value=clean(f.get("notes", "")))
        ws.cell(row=r, column=7, value=clean((f.get("source_commit") or "")[:10]))
    wrap_column(ws, 2)
    wrap_column(ws, 6)
    autosize(ws, max_width=70)


def tab_ci_commits(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("CI Commits")
    headers = ["SHA", "Date", "Message", "Root cause", "Checklist would catch?",
               "Post ready-for-review?"]
    write_header(ws, headers)
    commits = [c for c in (data.get("commits") or []) if c.get("is_ci_fix")]
    if not commits:
        add_no_data(ws, headers)
        autosize(ws)
        return
    for r, c in enumerate(commits, start=2):
        ws.cell(row=r, column=1, value=clean((c.get("sha") or "")[:10]))
        ws.cell(row=r, column=2, value=clean(c.get("date", "")))
        ws.cell(row=r, column=3, value=clean(c.get("message", "")))
        ws.cell(row=r, column=4, value=clean(c.get("ci_root_cause", "")))
        ws.cell(row=r, column=5, value=clean(c.get("checklist_would_catch", "")))
        ws.cell(row=r, column=6, value="Y" if c.get("post_ready_for_review") else "")
    wrap_column(ws, 3)
    wrap_column(ws, 4)
    autosize(ws, max_width=70)


def tab_docs(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("PR-Prep Docs")
    headers = ["Path", "Bytes", "Last modified", "Reviewer-useful?",
               "Keep under orchestrator model?", "Orphan?", "Notes"]
    write_header(ws, headers)
    files = (data.get("docs_audit") or {}).get("files") or []
    if not files:
        add_no_data(ws, headers)
        autosize(ws)
        return
    for r, f in enumerate(files, start=2):
        ws.cell(row=r, column=1, value=clean(f.get("path", "")))
        ws.cell(row=r, column=2, value=clean(f.get("bytes", 0)))
        ws.cell(row=r, column=3, value=clean(f.get("mtime", "")))
        ws.cell(row=r, column=4, value=clean(f.get("reviewer_useful", "")))
        ws.cell(row=r, column=5, value=clean(f.get("keep", "")))
        ws.cell(row=r, column=6, value="Y" if f.get("orphan") else "")
        ws.cell(row=r, column=7, value=clean(f.get("notes", "")))
    wrap_column(ws, 7)
    autosize(ws, max_width=70)


def tab_specs(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Specs Written")
    headers = ["Path", "Created", "Target PR", "Signals-ARCH-retro?", "Summary"]
    write_header(ws, headers)
    files = (data.get("specs") or {}).get("files") or []
    if not files:
        add_no_data(ws, headers)
        autosize(ws)
        return
    for r, f in enumerate(files, start=2):
        ws.cell(row=r, column=1, value=clean(f.get("path", "")))
        ws.cell(row=r, column=2, value=clean(f.get("created", "")))
        ws.cell(row=r, column=3, value=clean(f.get("target_pr", "")))
        ws.cell(row=r, column=4, value="Y" if f.get("retro_signal") else "")
        ws.cell(row=r, column=5, value=clean(f.get("summary", "")))
    wrap_column(ws, 5)
    autosize(ws, max_width=70)


def tab_sessions(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Sessions")
    headers = ["Session id", "First ts", "Last ts", "Duration (h)", "Msgs",
               "Compacted?", "Compaction events", "User prompts", "Near-done prompts",
               "Direction changes", "Branches seen", "Opening prompt", "Top tools"]
    write_header(ws, headers)
    sessions = (data.get("sessions") or {}).get("sessions") or []
    if not sessions:
        add_no_data(ws, headers)
        autosize(ws)
        return
    for r, s in enumerate(sessions, start=2):
        ws.cell(row=r, column=1, value=clean((s.get("id") or "")[:12]))
        ws.cell(row=r, column=2, value=clean(s.get("first_ts", "")))
        ws.cell(row=r, column=3, value=clean(s.get("last_ts", "")))
        ws.cell(row=r, column=4, value=clean(s.get("duration_hours", 0)))
        ws.cell(row=r, column=5, value=clean(s.get("message_count", 0)))
        ws.cell(row=r, column=6, value="Y" if s.get("compacted") else "")
        ws.cell(row=r, column=7, value=clean(s.get("compaction_events", 0)))
        ws.cell(row=r, column=8, value=clean(s.get("user_prompt_count", 0)))
        ws.cell(row=r, column=9, value=clean(s.get("near_done_prompts", 0)))
        ws.cell(row=r, column=10, value=clean(s.get("direction_changes", 0)))
        ws.cell(row=r, column=11, value=clean(", ".join(s.get("branches_seen")) or []))
        ws.cell(row=r, column=12, value=clean(s.get("summary") or s.get("opening_prompt", "")))
        tools = s.get("tools_used") or {}
        top = sorted(tools.items(), key=lambda kv: kv[1], reverse=True)[:5]
        ws.cell(row=r, column=13, value=clean(", ".join(f"{k}:{v}" for k, v in top)))
    wrap_column(ws, 12)
    autosize(ws, max_width=70)


def tab_files_touched(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Files Touched")
    headers = ["File", "Commits", "Additions", "Deletions"]
    write_header(ws, headers)
    counter: dict[str, dict[str, int]] = {}
    for c in data.get("commits") or []:
        for fp in c.get("files") or []:
            entry = counter.setdefault(fp, {"commits": 0, "additions": 0, "deletions": 0})
            entry["commits"] += 1
            entry["additions"] += c.get("additions", 0)
            entry["deletions"] += c.get("deletions", 0)
    rows = sorted(counter.items(), key=lambda kv: kv[1]["commits"], reverse=True)[:50]
    if not rows:
        add_no_data(ws, headers)
        autosize(ws)
        return
    for r, (fp, stats) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=clean(fp))
        ws.cell(row=r, column=2, value=clean(stats["commits"]))
        ws.cell(row=r, column=3, value=clean(stats["additions"]))
        ws.cell(row=r, column=4, value=clean(stats["deletions"]))
    autosize(ws, max_width=70)


def tab_friction(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Friction")
    headers = ["Signal", "Measurement", "Calendar-time cost", "Cognitive-load cost",
               "Evidence", "Notes"]
    write_header(ws, headers)
    items = data.get("friction") or []
    if not items:
        add_no_data(ws, headers)
        autosize(ws)
        return
    for r, it in enumerate(items, start=2):
        ws.cell(row=r, column=1, value=clean(it.get("signal", "")))
        ws.cell(row=r, column=2, value=clean(it.get("measurement", "")))
        ws.cell(row=r, column=3, value=clean(it.get("calendar_cost", "")))
        ws.cell(row=r, column=4, value=clean(it.get("cognitive_cost", "")))
        ws.cell(row=r, column=5, value=clean(it.get("evidence", "")))
        ws.cell(row=r, column=6, value=clean(it.get("notes", "")))
    wrap_column(ws, 5)
    wrap_column(ws, 6)
    autosize(ws, max_width=70)


def tab_sessions_commits(wb: Workbook, data: dict) -> None:
    """Join sessions to commits whose author-date falls within each session window."""
    ws = wb.create_sheet("Sessions → Commits")
    headers = ["Session id", "Window", "Duration (h)", "Msgs", "Compacted?",
               "Commits in window", "SHAs", "Dominant areas", "Opening prompt"]
    write_header(ws, headers)
    sessions = (data.get("sessions") or {}).get("sessions") or []
    commits = data.get("commits") or []
    if not sessions or not commits:
        add_no_data(ws, headers)
        autosize(ws)
        return

    # Pre-parse commit timestamps once.
    parsed: list[tuple[datetime | None, dict]] = [(parse_dt(c.get("date")), c) for c in commits]

    for r, s in enumerate(sessions, start=2):
        first = parse_dt(s.get("first_ts"))
        last = parse_dt(s.get("last_ts"))
        in_window: list[dict] = []
        if first and last:
            for cdt, c in parsed:
                if cdt and first <= cdt <= last:
                    in_window.append(c)
        shas_list = [(c.get("sha") or "")[:8] for c in in_window[:10]]
        shas = ", ".join(shas_list)
        if len(in_window) > 10:
            shas += f" +{len(in_window) - 10} more"

        area_count: dict[str, int] = {}
        for c in in_window:
            for fp in c.get("files") or []:
                prefix = fp.split("/", 1)[0] if "/" in fp else fp
                area_count[prefix] = area_count.get(prefix, 0) + 1
        top_areas = sorted(area_count.items(), key=lambda kv: -kv[1])[:3]
        areas = ", ".join(f"{a}({n})" for a, n in top_areas)

        window_str = (
            f"{first.strftime('%Y-%m-%d %H:%M') if first else '?'} → "
            f"{last.strftime('%Y-%m-%d %H:%M') if last else '?'}"
        )
        ws.cell(row=r, column=1, value=clean((s.get("id") or "")[:12]))
        ws.cell(row=r, column=2, value=clean(window_str))
        ws.cell(row=r, column=3, value=clean(s.get("duration_hours", 0)))
        ws.cell(row=r, column=4, value=clean(s.get("message_count", 0)))
        ws.cell(row=r, column=5, value="Y" if s.get("compacted") else "")
        ws.cell(row=r, column=6, value=len(in_window))
        ws.cell(row=r, column=7, value=clean(shas))
        ws.cell(row=r, column=8, value=clean(areas))
        ws.cell(row=r, column=9, value=clean(s.get("summary") or s.get("opening_prompt", "")))
    wrap_column(ws, 7)
    wrap_column(ws, 9)
    autosize(ws, max_width=60)


def tab_gantt_sessions(wb: Workbook, args: argparse.Namespace, data: dict) -> None:
    """Visual Gantt: rows = sessions, cols = days in window, filled bar = active."""
    ws = wb.create_sheet("Gantt (Sessions)")
    sessions = (data.get("sessions") or {}).get("sessions") or []
    if not sessions:
        ws.cell(row=1, column=1, value="(no sessions in window)")
        return

    start = parse_dt(args.window[0] + "T00:00:00+00:00")
    end = parse_dt(args.window[1] + "T23:59:59+00:00")
    if not start or not end:
        ws.cell(row=1, column=1, value="(invalid window)")
        return
    days = (end.date() - start.date()).days + 1

    headers = ["Session id", "Msgs", "Compact?", "Summary"] + [
        (start.date() + timedelta(days=i)).strftime("%m/%d")
        for i in range(days)
    ]
    write_header(ws, headers)

    # Color ramp by message count (quintiles across sessions).
    counts = sorted(s.get("message_count", 0) for s in sessions)
    def band_color(n: int) -> str:
        if not counts:
            return "B6D7A8"
        if n < counts[len(counts) // 5 or 1]: return "D9EAD3"  # pale green
        if n < counts[2 * len(counts) // 5 or 1]: return "B6D7A8"
        if n < counts[3 * len(counts) // 5 or 1]: return "93C47D"
        if n < counts[4 * len(counts) // 5 or 1]: return "6AA84F"
        return "38761D"  # deep green

    # Sort sessions by first_ts for readability.
    sorted_sessions = sorted(
        sessions,
        key=lambda s: (parse_dt(s.get("first_ts")) or datetime.max.replace(tzinfo=start.tzinfo)),
    )

    for r, s in enumerate(sorted_sessions, start=2):
        msgs = s.get("message_count", 0)
        ws.cell(row=r, column=1, value=clean((s.get("id") or "")[:12]))
        ws.cell(row=r, column=2, value=msgs)
        ws.cell(row=r, column=3, value="Y" if s.get("compacted") else "")
        ws.cell(row=r, column=4, value=clean(s.get("summary") or (s.get("opening_prompt") or "")[:80]))

        first = parse_dt(s.get("first_ts"))
        last = parse_dt(s.get("last_ts"))
        if not first or not last:
            continue
        color = band_color(msgs)
        fill = PatternFill("solid", fgColor=color)
        for d in range(days):
            day = start.date() + timedelta(days=d)
            day_start = datetime.combine(day, datetime.min.time(), tzinfo=start.tzinfo)
            day_end = datetime.combine(day, datetime.max.time(), tzinfo=start.tzinfo)
            # Session active on this day if its window overlaps [day_start, day_end].
            if first <= day_end and last >= day_start:
                ws.cell(row=r, column=5 + d, value="").fill = fill
    wrap_column(ws, 4)
    # Freeze first 4 metadata columns + header row.
    ws.freeze_panes = "E2"
    # Narrow day columns.
    for i in range(days):
        ws.column_dimensions[get_column_letter(5 + i)].width = 4
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 7
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 60


def tab_gantt_work(wb: Workbook, args: argparse.Namespace, data: dict) -> None:
    """Heatmap: rows = commit type, cols = days, cell = count (color intensity)."""
    ws = wb.create_sheet("Gantt (Work Types)")
    commits = data.get("commits") or []
    if not commits:
        ws.cell(row=1, column=1, value="(no commits)")
        return

    start = parse_dt(args.window[0] + "T00:00:00+00:00")
    end = parse_dt(args.window[1] + "T23:59:59+00:00")
    if not start or not end:
        ws.cell(row=1, column=1, value="(invalid window)")
        return
    days = (end.date() - start.date()).days + 1

    # Canonicalize type: ci fixes collapse to "fix(ci)".
    def commit_bucket(c: dict) -> str:
        msg = c.get("message", "") or ""
        if _CI_FIX_RE.match(msg):
            return "fix(ci)"
        t = (c.get("type") or "other").lower()
        return t if t else "other"

    headers = ["Work type", "Count"] + [
        (start.date() + timedelta(days=i)).strftime("%m/%d")
        for i in range(days)
    ]
    write_header(ws, headers)

    # Aggregate counts per (bucket, day).
    grid: dict[str, list[int]] = {}
    for c in commits:
        cdt = parse_dt(c.get("date"))
        if not cdt:
            continue
        d = (cdt.date() - start.date()).days
        if d < 0 or d >= days:
            continue
        bucket = commit_bucket(c)
        grid.setdefault(bucket, [0] * days)[d] += 1

    if not grid:
        ws.cell(row=2, column=1, value="(no dated commits)")
        return

    # Intensity ramp by count.
    def heat_color(n: int) -> str | None:
        if n <= 0: return None
        if n == 1: return "CFE2F3"
        if n == 2: return "9FC5E8"
        if n == 3: return "6FA8DC"
        if n == 4: return "3D85C6"
        return "0B5394"

    # Preferred row order.
    order = ["feat", "fix", "fix(ci)", "test", "refactor", "docs", "chore", "other"]
    rows_sorted = sorted(
        grid.items(),
        key=lambda kv: (order.index(kv[0]) if kv[0] in order else len(order), kv[0]),
    )

    for r, (bucket, counts) in enumerate(rows_sorted, start=2):
        total = sum(counts)
        ws.cell(row=r, column=1, value=clean(bucket)).font = Font(bold=True)
        ws.cell(row=r, column=2, value=total)
        for d, n in enumerate(counts):
            cell = ws.cell(row=r, column=3 + d, value=n if n else None)
            color = heat_color(n)
            if color:
                cell.fill = PatternFill("solid", fgColor=color)
                cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "C2"
    for i in range(days):
        ws.column_dimensions[get_column_letter(3 + i)].width = 4
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 7


def tab_refinements(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Refinements")
    headers = ["Letter", "Name", "Description", "Metric",
               "Current value", "Target on next PR", "Priority"]
    write_header(ws, headers)
    items = data.get("refinements") or []
    if not items:
        add_no_data(ws, headers)
        autosize(ws)
        return
    for r, it in enumerate(items, start=2):
        ws.cell(row=r, column=1, value=clean(it.get("letter", "")))
        ws.cell(row=r, column=2, value=clean(it.get("name", "")))
        ws.cell(row=r, column=3, value=clean(it.get("description", "")))
        ws.cell(row=r, column=4, value=clean(it.get("metric", "")))
        ws.cell(row=r, column=5, value=clean(it.get("current", "")))
        ws.cell(row=r, column=6, value=clean(it.get("target", "")))
        ws.cell(row=r, column=7, value=clean(it.get("priority", "")))
    wrap_column(ws, 3)
    wrap_column(ws, 4)
    autosize(ws, max_width=70)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--pr", required=True)
    ap.add_argument("--title", default="")
    ap.add_argument("--base", default="main")
    ap.add_argument("--head", required=True)
    ap.add_argument("--window", nargs=2, required=True, metavar=("START", "END"))
    ap.add_argument("--commits", help="commits.json from git-history agent")
    ap.add_argument("--docs-audit", help="docs-audit.json from docs agent")
    ap.add_argument("--ci-timeline", help="ci-timeline.json from CI agent")
    ap.add_argument("--sessions", help="sessions.json from analyze_sessions.py")
    ap.add_argument("--walkthrough-fixes", help="walkthrough-fixes.json")
    ap.add_argument("--specs", help="specs.json")
    ap.add_argument("--friction", help="friction.json")
    ap.add_argument("--refinements", help="refinements.json")
    ap.add_argument("--ready-for-review-ts",
                    help="ISO timestamp the PR went ready-for-review (overrides inference)")
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    data = {
        "commits": load(args.commits) or [],
        "docs_audit": load(args.docs_audit) or {},
        "ci_timeline": load(args.ci_timeline) or {},
        "sessions": load(args.sessions) or {},
        "walkthrough_fixes": load(args.walkthrough_fixes) or [],
        "specs": load(args.specs) or {},
        "friction": load(args.friction) or [],
        "refinements": load(args.refinements) or [],
    }

    # Enrich commits with inferred flags (respects pre-existing manual values).
    data["commits"], rfr_dt = enrich_commits(
        data["commits"],
        data["walkthrough_fixes"],
        data["ci_timeline"],
        args.ready_for_review_ts,
    )
    # Surface the resolved ready-for-review into the ci_timeline view for Summary tab.
    if rfr_dt and not data["ci_timeline"].get("ready_for_review_ts"):
        data["ci_timeline"]["ready_for_review_ts"] = rfr_dt.isoformat()

    wb = Workbook()
    # Drop default sheet
    default = wb.active
    wb.remove(default)

    tab_summary(wb, args, data)
    tab_gantt_sessions(wb, args, data)
    tab_gantt_work(wb, args, data)
    tab_timeline(wb, data)
    tab_commits(wb, data)
    tab_walkthrough_fixes(wb, data)
    tab_ci_commits(wb, data)
    tab_docs(wb, data)
    tab_specs(wb, data)
    tab_sessions(wb, data)
    tab_sessions_commits(wb, data)
    tab_files_touched(wb, data)
    tab_friction(wb, data)
    tab_refinements(wb, data)

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"Wrote {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
